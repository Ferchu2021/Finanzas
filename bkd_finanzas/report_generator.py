from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date, datetime
from typing import List, Dict
from io import BytesIO
import models


def generar_reporte_pdf_gastos(gastos: List[models.Gasto], fecha_inicio: date, fecha_fin: date) -> BytesIO:
    """Genera un reporte PDF de gastos por rango de fechas"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    story = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#2c3e50'),
        spaceAfter=30,
        alignment=1  # Centrado
    )
    
    # Título
    story.append(Paragraph("Reporte de Gastos", title_style))
    story.append(Spacer(1, 0.2*inch))
    
    # Información del período
    periodo_text = f"Período: {fecha_inicio.strftime('%d/%m/%Y')} - {fecha_fin.strftime('%d/%m/%Y')}"
    story.append(Paragraph(periodo_text, styles['Normal']))
    story.append(Paragraph(f"Fecha de generación: {date.today().strftime('%d/%m/%Y')}", styles['Normal']))
    story.append(Spacer(1, 0.3*inch))
    
    # Calcular totales
    total_ars = sum(g.monto for g in gastos if g.moneda == models.TipoMoneda.PESOS)
    total_usd = sum(g.monto for g in gastos if g.moneda == models.TipoMoneda.DOLARES)
    cantidad_gastos = len(gastos)
    
    # Resumen
    resumen_data = [
        ['Total de Gastos', f'{cantidad_gastos}'],
        ['Total en ARS', f'${total_ars:,.2f}'],
        ['Total en USD', f'${total_usd:,.2f}']
    ]
    resumen_table = Table(resumen_data, colWidths=[3*inch, 2*inch])
    resumen_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#f8f9fa')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#2c3e50')),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('TOPPADDING', (0, 0), (-1, -1), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#dee2e6'))
    ]))
    story.append(resumen_table)
    story.append(Spacer(1, 0.3*inch))
    
    # Tabla de gastos
    if gastos:
        # Ordenar por fecha
        gastos_ordenados = sorted(gastos, key=lambda x: x.fecha)
        
        # Encabezados
        data = [['Fecha', 'Descripción', 'Tipo', 'Categoría', 'Monto', 'Moneda']]
        
        # Datos
        for gasto in gastos_ordenados:
            descripcion = gasto.descripcion or 'Sin descripción'
            categoria = gasto.categoria or '-'
            data.append([
                gasto.fecha.strftime('%d/%m/%Y'),
                descripcion[:50],  # Limitar longitud
                gasto.tipo.value,
                categoria,
                f'${gasto.monto:,.2f}',
                gasto.moneda.value
            ])
        
        # Crear tabla
        table = Table(data, colWidths=[0.8*inch, 2.5*inch, 1*inch, 1*inch, 1*inch, 0.7*inch])
        table.setStyle(TableStyle([
            # Encabezado
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498db')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TOPPADDING', (0, 0), (-1, 0), 12),
            # Filas de datos
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.HexColor('#2c3e50')),
            ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
            ('ALIGN', (4, 1), (4, -1), 'RIGHT'),  # Monto alineado a la derecha
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#dee2e6')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        story.append(table)
    else:
        story.append(Paragraph("No hay gastos registrados en el período seleccionado.", styles['Normal']))
    
    # Generar PDF
    doc.build(story)
    buffer.seek(0)
    return buffer


def generar_reporte_excel_gastos(gastos: List[models.Gasto], fecha_inicio: date, fecha_fin: date) -> BytesIO:
    """Genera un reporte Excel de gastos por rango de fechas"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de Gastos"
    
    # Estilos
    header_fill = PatternFill(start_color="3498db", end_color="3498db", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=14)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Título
    ws['A1'] = "Reporte de Gastos"
    ws['A1'].font = title_font
    ws.merge_cells('A1:F1')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Período
    ws['A2'] = f"Período: {fecha_inicio.strftime('%d/%m/%Y')} - {fecha_fin.strftime('%d/%m/%Y')}"
    ws['A3'] = f"Fecha de generación: {date.today().strftime('%d/%m/%Y')}"
    
    # Resumen
    row = 5
    ws[f'A{row}'] = "Resumen"
    ws[f'A{row}'].font = Font(bold=True, size=11)
    row += 1
    
    total_ars = sum(g.monto for g in gastos if g.moneda == models.TipoMoneda.PESOS)
    total_usd = sum(g.monto for g in gastos if g.moneda == models.TipoMoneda.DOLARES)
    cantidad_gastos = len(gastos)
    
    ws[f'A{row}'] = "Total de Gastos:"
    ws[f'B{row}'] = cantidad_gastos
    row += 1
    ws[f'A{row}'] = "Total en ARS:"
    ws[f'B{row}'] = total_ars
    ws[f'B{row}'].number_format = '"$"#,##0.00'
    row += 1
    ws[f'A{row}'] = "Total en USD:"
    ws[f'B{row}'] = total_usd
    ws[f'B{row}'].number_format = '"$"#,##0.00'
    row += 2
    
    # Encabezados de tabla
    headers = ['Fecha', 'Descripción', 'Tipo', 'Categoría', 'Monto', 'Moneda']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    row += 1
    
    # Datos
    if gastos:
        gastos_ordenados = sorted(gastos, key=lambda x: x.fecha)
        for gasto in gastos_ordenados:
            ws.cell(row=row, column=1, value=gasto.fecha.strftime('%d/%m/%Y')).border = border
            ws.cell(row=row, column=2, value=gasto.descripcion or 'Sin descripción').border = border
            ws.cell(row=row, column=3, value=gasto.tipo.value).border = border
            ws.cell(row=row, column=4, value=gasto.categoria or '-').border = border
            monto_cell = ws.cell(row=row, column=5, value=gasto.monto)
            monto_cell.number_format = '"$"#,##0.00'
            monto_cell.alignment = Alignment(horizontal='right')
            monto_cell.border = border
            ws.cell(row=row, column=6, value=gasto.moneda.value).border = border
            row += 1
        
        # Fila de totales
        ws.cell(row=row, column=4, value="TOTAL:").font = Font(bold=True)
        ws.cell(row=row, column=4).alignment = Alignment(horizontal='right')
        ws.cell(row=row, column=5, value=total_ars + total_usd).font = Font(bold=True)
        ws.cell(row=row, column=5).number_format = '"$"#,##0.00'
        ws.cell(row=row, column=5).alignment = Alignment(horizontal='right')
    else:
        ws.merge_cells(f'A{row}:F{row}')
        ws.cell(row=row, column=1, value="No hay gastos registrados en el período seleccionado.")
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
    
    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 10
    
    # Guardar en buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


